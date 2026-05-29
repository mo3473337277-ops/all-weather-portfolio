"""Excel 多 sheet 报告导出（openpyxl）。

输出 output/report.xlsx，包含 11 个 sheet：
1. 推荐总览  2. 核心指标  3. 分年化收益  4. 风险贡献  5. 宏观情景
6. 关键事件  7. 滚动统计  8. Bootstrap   9. 持仓清单 10. 净值曲线 11. 权重明细

格式约定：
- 百分比列 0.00%，小数列 0.0000
- 表头加粗 + 浅蓝填充
- 负百分比红色字体
- 列宽自适应内容
"""
from pathlib import Path
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from .config import BUCKET_GROUPS, ETF_META, OUTPUT_DIR
from .config import PORTFOLIO_TAGS

# === 样式常量 ===
HEADER_FILL = PatternFill("solid", fgColor="DCE6F1")
TITLE_FILL = PatternFill("solid", fgColor="4F81BD")
RECO_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(bold=True, color="000000")
TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
NEG_FONT = Font(color="C00000")
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

PCT_FMT = "0.00%"
NUM_FMT = "0.0000"
INT_FMT = "#,##0"


def _write_title(ws, row, text, n_cols):
    """写一个跨列的标题行。"""
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    c = ws.cell(row=row, column=1)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = CENTER
    ws.row_dimensions[row].height = 22


def _write_headers(ws, row, headers):
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER


def _autofit(ws, min_width=10, max_width=40):
    """根据内容长度估列宽（中文 1 字 ≈ 2 单位）。"""
    for col_idx in range(1, ws.max_column + 1):
        max_len = min_width
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            length = sum(2 if ord(ch) > 127 else 1 for ch in s)
            max_len = max(max_len, length + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, max_width)


def _apply_pct_neg_red(ws, first_row, last_row, first_col, last_col):
    """对一个区域：百分比格式 + 负数红字。"""
    rng = f"{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}"
    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.number_format = PCT_FMT
            cell.alignment = RIGHT
            cell.border = BORDER
    rule = CellIsRule(operator="lessThan", formula=["0"], font=NEG_FONT)
    ws.conditional_formatting.add(rng, rule)


# ============================================================
# 各 sheet 写入函数
# ============================================================

def _sheet_recommendation(wb):
    ws = wb.create_sheet("推荐总览")
    _write_title(ws, 1, "桥水全天候策略 · 中国版 - 方案推荐", 4)
    _write_headers(ws, 3, ["推荐度", "方案", "标签", "说明"])
    notes = {
        "V3c 多元": "海外 8% 国别分散最强，综合 Sharpe 最高，蒙特卡洛下沿最厚实",
        "V3b 平衡": "长债 58%，回撤最浅 -2.26%，求稳者首选",
        "V3d 商品偏重": "商品 27%，CAGR 最高，抗滞胀但回撤更深",
    }
    # PORTFOLIO_TAGS 的顺序不是推荐顺序，按 stars 数量重排
    items = sorted(PORTFOLIO_TAGS.items(), key=lambda kv: -len(kv[1]["stars"]))
    for i, (port, tag) in enumerate(items, start=4):
        ws.cell(row=i, column=1, value=tag["stars"]).alignment = CENTER
        ws.cell(row=i, column=2, value=port).alignment = LEFT
        ws.cell(row=i, column=3, value=tag["label"]).alignment = LEFT
        ws.cell(row=i, column=4, value=notes.get(port, "")).alignment = LEFT
        if "★★★" in tag["stars"]:
            for c in range(1, 5):
                ws.cell(row=i, column=c).fill = RECO_FILL
        for c in range(1, 5):
            ws.cell(row=i, column=c).border = BORDER

    # 注脚
    ws.cell(row=8, column=1, value="📌 默认主推 V3c 多元；其它两套是「特殊偏好」备选。")
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=4)
    ws.cell(row=8, column=1).alignment = LEFT
    _autofit(ws)


def _sheet_perf(wb, perf_results):
    ws = wb.create_sheet("核心指标")
    _write_title(ws, 1, "三策略 × 三档现金 核心收益指标", 10)
    headers = ["方案", "档位", "累计收益", "CAGR", "波动", "最大回撤", "Sharpe", "Calmar", "期末净值", "D_excess"]
    _write_headers(ws, 3, headers)
    pct_cols = {3, 4, 5, 6, 10}  # 累计/CAGR/波动/MDD/D_excess
    num_cols = {7, 8, 9}         # Sharpe/Calmar/final_nv

    r = 4
    last_port = None
    for (port, tier), m in perf_results.items():
        if last_port and port != last_port:
            r += 1  # 空一行分组
        last_port = port
        ws.cell(row=r, column=1, value=port).alignment = LEFT
        ws.cell(row=r, column=2, value=tier).alignment = CENTER
        ws.cell(row=r, column=3, value=m["cum_return"])
        ws.cell(row=r, column=4, value=m["cagr"])
        ws.cell(row=r, column=5, value=m["vol"])
        ws.cell(row=r, column=6, value=m["mdd"])
        ws.cell(row=r, column=7, value=m["sharpe"])
        ws.cell(row=r, column=8, value=m["calmar"])
        ws.cell(row=r, column=9, value=m["final_nv"])
        ws.cell(row=r, column=10, value=m["geometric_excess_d"])
        for c in pct_cols:
            ws.cell(row=r, column=c).number_format = PCT_FMT
            ws.cell(row=r, column=c).alignment = RIGHT
        for c in num_cols:
            ws.cell(row=r, column=c).number_format = NUM_FMT
            ws.cell(row=r, column=c).alignment = RIGHT
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = BORDER
        # 主推 V3c 100% RP 高亮
        if port == "V3c 多元" and tier == "100% RP":
            for c in range(1, 11):
                ws.cell(row=r, column=c).fill = RECO_FILL
        r += 1
    # 负数红字（MDD 一定是负，累计也可能负）
    rng = f"C4:J{r-1}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], font=NEG_FONT))
    _autofit(ws)


def _sheet_yearly(wb, yearly_results, years=None):
    ws = wb.create_sheet("分年化收益")
    if years is None:
        years = list(range(2008, 2026))
    _write_title(ws, 1, f"分年化收益（100% RP 档）", len(years) + 1)
    _write_headers(ws, 3, ["方案"] + [str(y) for y in years])
    for i, (port, s) in enumerate(yearly_results.items(), start=4):
        ws.cell(row=i, column=1, value=port).alignment = LEFT
        for j, y in enumerate(years, start=2):
            v = s.get(y, None)
            ws.cell(row=i, column=j, value=float(v) if v is not None and pd.notna(v) else None)
    _apply_pct_neg_red(ws, 4, 3 + len(yearly_results), 2, 1 + len(years))
    for r in range(4, 4 + len(yearly_results)):
        ws.cell(row=r, column=1).border = BORDER
    _autofit(ws)


def _sheet_risk_contrib(wb, rc_results):
    if not rc_results:
        return
    ws = wb.create_sheet("风险贡献")
    ports = list(rc_results.keys())
    _write_title(ws, 1, "桶级风险贡献分解（协方差视角）", len(ports) + 1)
    _write_headers(ws, 3, ["桶"] + ports)
    buckets = list(rc_results[ports[0]].keys())
    for i, b in enumerate(buckets, start=4):
        ws.cell(row=i, column=1, value=b).alignment = LEFT
        for j, p in enumerate(ports, start=2):
            ws.cell(row=i, column=j, value=float(rc_results[p][b]))
    _apply_pct_neg_red(ws, 4, 3 + len(buckets), 2, 1 + len(ports))
    for r in range(4, 4 + len(buckets)):
        ws.cell(row=r, column=1).border = BORDER
    _autofit(ws)


def _sheet_regime(wb, regime_results):
    ws = wb.create_sheet("宏观情景")
    ports = list(regime_results.keys())
    regimes = ["股牛+债牛", "股牛+债熊", "股熊+债牛", "股熊+债熊"]
    _write_title(ws, 1, "4 宏观情景平均季度收益（100% RP 档）", len(regimes) + 1)
    first = regime_results[ports[0]]
    headers = ["方案"] + [f"{r}（n={first[r]['n']}）" for r in regimes]
    _write_headers(ws, 3, headers)
    for i, p in enumerate(ports, start=4):
        ws.cell(row=i, column=1, value=p).alignment = LEFT
        for j, r_label in enumerate(regimes, start=2):
            ws.cell(row=i, column=j, value=float(regime_results[p][r_label]["avg"]))
    _apply_pct_neg_red(ws, 4, 3 + len(ports), 2, 1 + len(regimes))
    for r in range(4, 4 + len(ports)):
        ws.cell(row=r, column=1).border = BORDER
    _autofit(ws)


def _sheet_events(wb, event_results):
    ws = wb.create_sheet("关键事件")
    ports = list(event_results.keys())
    events = list(event_results[ports[0]].keys())
    _write_title(ws, 1, "关键事件期累计收益（100% RP 档）", len(ports) + 1)
    _write_headers(ws, 3, ["事件"] + ports)
    for i, ev in enumerate(events, start=4):
        ws.cell(row=i, column=1, value=ev).alignment = LEFT
        for j, p in enumerate(ports, start=2):
            v = event_results[p][ev]
            ws.cell(row=i, column=j, value=float(v) if pd.notna(v) else None)
    _apply_pct_neg_red(ws, 4, 3 + len(events), 2, 1 + len(ports))
    for r in range(4, 4 + len(events)):
        ws.cell(row=r, column=1).border = BORDER
    _autofit(ws)


def _sheet_rolling(wb, rolling_results):
    ws = wb.create_sheet("滚动统计")
    headers = ["方案", "年化-min", "年化-中位", "年化-max", "1y回撤-最差", "负收益年%"]
    _write_title(ws, 1, "滚动 1 年表现（100% RP 档）", len(headers))
    _write_headers(ws, 3, headers)
    for i, (port, s) in enumerate(rolling_results.items(), start=4):
        ws.cell(row=i, column=1, value=port).alignment = LEFT
        ws.cell(row=i, column=2, value=float(s["ann_min"]))
        ws.cell(row=i, column=3, value=float(s["ann_med"]))
        ws.cell(row=i, column=4, value=float(s["ann_max"]))
        ws.cell(row=i, column=5, value=float(s["dd_min"]))
        ws.cell(row=i, column=6, value=float(s["neg_year_pct"]))
    _apply_pct_neg_red(ws, 4, 3 + len(rolling_results), 2, 6)
    for r in range(4, 4 + len(rolling_results)):
        ws.cell(row=r, column=1).border = BORDER
    _autofit(ws)


def _sheet_bootstrap(wb, boot_results):
    ws = wb.create_sheet("Bootstrap")
    headers = ["方案", "5%分位", "25%分位", "中位数", "75%分位", "95%分位", "年化中位", "亏损概率"]
    _write_title(ws, 1, "Block Bootstrap 5 年期累计收益分布（1000 次模拟）", len(headers))
    _write_headers(ws, 3, headers)
    for i, (port, b) in enumerate(boot_results.items(), start=4):
        ws.cell(row=i, column=1, value=port).alignment = LEFT
        ws.cell(row=i, column=2, value=float(b["p05"]))
        ws.cell(row=i, column=3, value=float(b["p25"]))
        ws.cell(row=i, column=4, value=float(b["p50"]))
        ws.cell(row=i, column=5, value=float(b["p75"]))
        ws.cell(row=i, column=6, value=float(b["p95"]))
        ws.cell(row=i, column=7, value=float(b["ann_median"]))
        ws.cell(row=i, column=8, value=float(b["loss_prob"]))
    _apply_pct_neg_red(ws, 4, 3 + len(boot_results), 2, 8)
    for r in range(4, 4 + len(boot_results)):
        ws.cell(row=r, column=1).border = BORDER
    _autofit(ws)


def _sheet_holdings(wb, weights_dict, principal=1_000_000):
    ws = wb.create_sheet("持仓清单")
    ports = list(weights_dict.keys())
    _write_title(ws, 1, f"持仓清单（按 {principal:,.0f} 本金，100% RP 档）", 3 + len(ports))
    headers = ["桶", "资产", "代码"] + ports
    _write_headers(ws, 3, headers)
    r = 4
    for bk, lst in BUCKET_GROUPS.items():
        for asset in lst:
            meta = ETF_META[asset]
            ws.cell(row=r, column=1, value=bk).alignment = LEFT
            ws.cell(row=r, column=2, value=meta["name"]).alignment = LEFT
            ws.cell(row=r, column=3, value=meta["code"]).alignment = CENTER
            for j, p in enumerate(ports, start=4):
                amt = weights_dict[p].get(asset, 0) * principal
                cell = ws.cell(row=r, column=j, value=float(amt))
                cell.number_format = INT_FMT
                cell.alignment = RIGHT
            for c in range(1, 4 + len(ports)):
                ws.cell(row=r, column=c).border = BORDER
            r += 1
    # 合计行
    ws.cell(row=r, column=1, value="合计").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    for c in range(2, 4):
        ws.cell(row=r, column=c).fill = HEADER_FILL
    for j, p in enumerate(ports, start=4):
        cell = ws.cell(row=r, column=j, value=float(principal))
        cell.number_format = INT_FMT
        cell.alignment = RIGHT
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for c in range(1, 4 + len(ports)):
        ws.cell(row=r, column=c).border = BORDER
    _autofit(ws)


def _sheet_nv_curves(wb, nv_results):
    ws = wb.create_sheet("净值曲线")
    df = pd.DataFrame({f"{p}_{t}": nv for (p, t), nv in nv_results.items()})
    headers = ["日期"] + list(df.columns)
    _write_title(ws, 1, "9 条净值曲线（V3b/V3c/V3d × 100/85/70% RP）", len(headers))
    _write_headers(ws, 3, headers)
    # 数据量大，直接写
    for i, (idx, row) in enumerate(df.iterrows(), start=4):
        ws.cell(row=i, column=1, value=idx.strftime("%Y-%m-%d"))
        for j, col in enumerate(df.columns, start=2):
            cell = ws.cell(row=i, column=j, value=float(row[col]))
            cell.number_format = NUM_FMT
    _autofit(ws, max_width=14)
    ws.freeze_panes = "B4"


def _sheet_weights(wb, weights_dict):
    if not weights_dict:
        return
    ws = wb.create_sheet("权重明细")
    ports = list(weights_dict.keys())
    _write_title(ws, 1, "三策略权重明细", 1 + len(ports))
    _write_headers(ws, 3, ["资产"] + ports)
    assets = list(weights_dict[ports[0]].index)
    for i, a in enumerate(assets, start=4):
        ws.cell(row=i, column=1, value=a).alignment = LEFT
        for j, p in enumerate(ports, start=2):
            cell = ws.cell(row=i, column=j, value=float(weights_dict[p][a]))
            cell.number_format = PCT_FMT
            cell.alignment = RIGHT
        for c in range(1, 2 + len(ports)):
            ws.cell(row=i, column=c).border = BORDER
    # 合计
    r = 4 + len(assets)
    ws.cell(row=r, column=1, value="合计").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    for j, p in enumerate(ports, start=2):
        cell = ws.cell(row=r, column=j, value=float(weights_dict[p].sum()))
        cell.number_format = PCT_FMT
        cell.alignment = RIGHT
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for c in range(1, 2 + len(ports)):
        ws.cell(row=r, column=c).border = BORDER
    _autofit(ws)


# ============================================================
# 主入口
# ============================================================

def save_excel_report(
    nv_results,
    perf_results,
    yearly_results,
    rc_results,
    regime_results,
    event_results,
    rolling_results,
    boot_results,
    weights_dict,
    filename="report.xlsx",
):
    """生成 output/report.xlsx 多 sheet 综合报告。"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # 删默认 sheet
    wb.remove(wb.active)

    _sheet_recommendation(wb)
    _sheet_perf(wb, perf_results)
    _sheet_yearly(wb, yearly_results)
    _sheet_risk_contrib(wb, rc_results)
    _sheet_regime(wb, regime_results)
    _sheet_events(wb, event_results)
    _sheet_rolling(wb, rolling_results)
    _sheet_bootstrap(wb, boot_results)
    _sheet_holdings(wb, weights_dict)
    _sheet_nv_curves(wb, nv_results)
    _sheet_weights(wb, weights_dict)

    path = OUTPUT_DIR / filename
    wb.save(path)
    return path
